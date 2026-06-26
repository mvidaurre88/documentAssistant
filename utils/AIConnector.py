import io
import logging, mammoth, base64, anthropic, httpx, streamlit as st
import os
import time
import zipfile
from pathlib import Path
from openpyxl import load_workbook
from openai import OpenAI
import openai

logger = logging.getLogger(__name__)

# EXTENSIONES SOPORTADAS -----------------------------------------------------------------------------
SUPPORTED_EXTENSIONS = {".pdf", ".docx", ".txt", ".png", ".py", ".xlsm", ".xlsx", ".json", ".atmx", ".xml"}

# EJEMPLOS DE REFERENCIA -----------------------------------------------------------------------------
_EJEMPLOS_DIR = Path(__file__).parent / "ejemplos_pdd"  # carpeta con los .md de referencia


def _cargar_ejemplos() -> str:
    """Lee los .md de la carpeta de ejemplos y los envuelve en <ejemplos_referencia>."""
    if not _EJEMPLOS_DIR.exists():
        logger.warning(f"Carpeta de ejemplos no encontrada: {_EJEMPLOS_DIR}")
        return ""
    partes = ["<ejemplos_referencia>"]
    archivos = sorted(_EJEMPLOS_DIR.glob("*.md"))
    for i, p in enumerate(archivos, start=1):
        partes.append(f"### EJEMPLO {i} ({p.stem})\n{p.read_text(encoding='utf-8')}")
    partes.append("</ejemplos_referencia>")
    logger.info(f"Ejemplos de referencia cargados: {len(archivos)}")
    return "\n\n".join(partes) if archivos else ""


# WRAPPER PARA ARCHIVOS EXTRAÍDOS DE UN ZIP ----------------------------------------------------------
class _ZipFileWrapper:
    """Simula la interfaz de UploadedFile para archivos extraídos de un zip."""
    def __init__(self, name, data):
        self.name = name
        self._data = data

    def getvalue(self):
        return self._data


# ENCODE FILES TO SEND TO AI
def encode_file(file_or_path):
    """Acepta un Path o un objeto tipo UploadedFile de Streamlit.
    Devuelve un bloque (dict), una lista de bloques (para ZIPs), o None."""

    # Determinar si es Path o file-like
    if isinstance(file_or_path, Path):
        if not file_or_path.exists():
            logger.error(f"Archivo no encontrado: {file_or_path}")
            return None
        name = file_or_path.name
        ext = file_or_path.suffix.lower()
        read_bytes = lambda: file_or_path.read_bytes()
        read_text = lambda: file_or_path.read_text(encoding="utf-8")
        open_for_mammoth = lambda: open(file_or_path, "rb")
        path_for_openpyxl = file_or_path
    else:
        # UploadedFile de Streamlit u objeto similar
        name = file_or_path.name
        ext = os.path.splitext(name)[1].lower()
        read_bytes = lambda: file_or_path.getvalue()
        read_text = lambda: file_or_path.getvalue().decode("utf-8")
        open_for_mammoth = lambda: io.BytesIO(file_or_path.getvalue())
        path_for_openpyxl = io.BytesIO(file_or_path.getvalue())

    # ZIP: descomprimir y procesar recursivamente cada archivo soportado
    if ext == ".zip":
        blocks = []
        with zipfile.ZipFile(io.BytesIO(read_bytes())) as z:
            for info in z.infolist():
                if info.is_dir():
                    continue
                inner_ext = os.path.splitext(info.filename)[1].lower()
                if inner_ext not in SUPPORTED_EXTENSIONS and inner_ext != ".zip":
                    logger.info(f"Saltando archivo no soportado dentro del zip: {info.filename}")
                    continue
                with z.open(info) as inner_file:
                    inner_bytes = inner_file.read()
                wrapper = _ZipFileWrapper(info.filename, inner_bytes)
                result = encode_file(wrapper)
                if result is None:
                    continue
                if isinstance(result, list):
                    blocks.extend(result)
                else:
                    blocks.append(result)
        logger.info(f"ZIP procesado: {name} | Bloques extraídos: {len(blocks)}")
        return blocks

    if ext not in SUPPORTED_EXTENSIONS:
        logger.error(f"Tipo no soportado: '{ext}'")
        return None

    block = None

    if ext == ".png":
        data = base64.standard_b64encode(read_bytes()).decode("utf-8")
        block = {"type": "image", "source": {"type": "base64", "media_type": "image/png", "data": data}}

    elif ext == ".pdf":
        data = base64.standard_b64encode(read_bytes()).decode("utf-8")
        block = {"type": "document", "source": {"type": "base64", "media_type": "application/pdf", "data": data}}

    elif ext == ".docx":
        docx_bytes = read_bytes()
        return _process_docx(docx_bytes, name)
    
    elif ext in (".txt", ".py", ".json", ".atmx", ".xml"):
        block = {"type": "text", "text": f"[Contenido de {name}]\n{read_text()}"}

    elif ext in (".xlsm", ".xlsx"):
        wb = load_workbook(path_for_openpyxl, read_only=True, keep_vba=True)
        try:
            lines = []
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                lines.append(f"### Hoja: {sheet_name}")
                for row in ws.iter_rows(values_only=True):
                    row_text = "\t".join("" if cell is None else str(cell) for cell in row)
                    if row_text.strip():
                        lines.append(row_text)
        finally:
            wb.close()
        block = {"type": "text", "text": f"[Contenido de {name}]\n" + "\n".join(lines)}

    # Log de tokens
    if block is not None:
        if block["type"] == "text":
            estimated_tokens = len(block["text"]) // 4
        elif block["type"] == "image":
            estimated_tokens = len(block["source"]["data"]) // 4
        elif block["type"] == "document":
            estimated_tokens = len(block["source"]["data"]) // 4
        else:
            estimated_tokens = 0

        logger.info(
            f"Archivo procesado: {name} | "
            f"Tipo: {block['type']} | "
            f"Tokens estimados: ~{estimated_tokens:,}"
        )

    return block

TRANSIENT_ERRORS = (    anthropic.APIConnectionError, anthropic.APITimeoutError,
                        anthropic.InternalServerError, anthropic.RateLimitError,
                        openai.APIConnectionError, openai.APITimeoutError,
                        openai.InternalServerError, openai.RateLimitError,
                        httpx.RemoteProtocolError, httpx.ReadTimeout, httpx.ConnectTimeout)
API_ERRORS = (anthropic.APIError, openai.APIError)

# SEND TO AI
def _anthropic_to_openai_blocks(blocks: list) -> list:
    """Convierte bloques del formato Anthropic al de OpenAI/LM Studio."""
    out = []
    for b in blocks:
        if b["type"] == "text":
            out.append(b)  # idéntico en ambos
        elif b["type"] == "image":
            src = b["source"]
            uri = f"data:{src['media_type']};base64,{src['data']}"
            out.append({"type": "image_url", "image_url": {"url": uri}})
        else:
            # 'document'/PDF: qwen2.5-vl no lo soporta -> se omite
            logger.warning(f"Bloque '{b['type']}' omitido (no soportado por QWEN).")
    return out


def _stream_anthropic(client, model, content_blocks, max_tokens, system=None):
    kwargs = dict(
        model=model,
        max_tokens=max_tokens,
        messages=[{"role": "user", "content": content_blocks}],
    )
    if system:
        kwargs["system"] = system
    with client.messages.stream(**kwargs) as stream:
        yield from stream.text_stream


def _stream_openai(client, model, content_blocks, max_tokens, system=None):
    messages = []
    if system:
        # QWEN/LM Studio no cachea: aplanar los bloques de texto a un system role plano
        sys_text = "\n\n".join(b["text"] for b in system if b.get("type") == "text")
        if sys_text.strip():
            messages.append({"role": "system", "content": sys_text})
    messages.append({"role": "user", "content": content_blocks})
    stream = client.chat.completions.create(
        model=model,
        max_tokens=max_tokens,
        messages=messages,
        stream=True,
    )
    for chunk in stream:
        delta = chunk.choices[0].delta.content
        if delta:
            yield delta


def send_to_ai(prompt: str, files: list, maxTokens: int = 4096) -> str:

    model = st.secrets.get("MODEL")

    if model == "QWEN":
        key = st.secrets.get("API_KEY_QWEN")
        client = OpenAI(base_url="http://192.168.1.48:8000/v1", api_key=key)
        model_name = "qwen/qwen2.5-vl-7b"
        stream_fn = _stream_openai
        maxTokens = min(maxTokens, 4096)
    elif model == "CLAUDE":
        key = st.secrets.get("API_KEY_CLAUDE")
        client = anthropic.Anthropic(api_key=key, timeout=900.0)
        model_name = "claude-sonnet-4-6"
        stream_fn = _stream_anthropic
    else:
        logger.error(f"MODEL no reconocido: {model!r}")
        raise ValueError(f"MODEL no reconocido: {model!r}")

    if not key:
        logger.error("API_KEY no encontrada en las variables de entorno.")
        raise EnvironmentError("API_KEY no encontrada en las variables de entorno.")

    # VALIDAR PROMPT
    if not prompt or not prompt.strip():
        logger.error("El prompt recibido está vacío.")
        raise ValueError("El prompt está vacío.")
    prompt_text = prompt.strip()

    # ENCODE FILES (formato Anthropic)
    content_blocks = []
    for file in files:
        result = encode_file(file)
        if result is None:
            continue
        if isinstance(result, list):
            content_blocks.extend(result)
        else:
            content_blocks.append(result)

    if not content_blocks and files:
        logger.error("No se pudieron procesar los archivos proporcionados.")
        raise ValueError("Ningún archivo pudo ser procesado.")

    # Prompt al final
    content_blocks.append({"type": "text", "text": prompt_text})

    # Si el backend es OpenAI/LM Studio, adaptar el formato de los bloques
    if model == "QWEN":
        content_blocks = _anthropic_to_openai_blocks(content_blocks)

    # SYSTEM: ejemplos de referencia (cacheados en Claude, system plano en QWEN)
    doc_type = st.session_state.get("doc_type")
    system = None
    if doc_type == "PDD":
        ejemplos = _cargar_ejemplos()
        if ejemplos:
            if model == "CLAUDE":
                system = [{
                    "type": "text",
                    "text": ejemplos,
                    "cache_control": {"type": "ephemeral"},
                }]
            else:
                system = [{"type": "text", "text": ejemplos}]

    # RETRY con backoff
    max_retries = 3
    last_error = None
    progress_placeholder = st.empty()

    for attempt in range(max_retries):
        try:
            full_response = ""
            for text in stream_fn(client, model_name, content_blocks, maxTokens, system):
                full_response += text
                if len(full_response) % 200 < len(text):
                    progress_placeholder.caption(
                        f"⚙️ Generando respuesta... {len(full_response):,} caracteres"
                    )
            progress_placeholder.empty()
            logger.info(
                f"Respuesta IA generada exitosamente "
                f"(intento {attempt + 1}, {len(full_response)} chars)"
            )
            return full_response

        except TRANSIENT_ERRORS as e:
            last_error = e
            logger.warning(
                f"Error transitorio en intento {attempt + 1}/{max_retries}: "
                f"{type(e).__name__}: {e}"
            )
            if attempt < max_retries - 1:
                wait = 2 ** attempt
                logger.info(f"Reintentando en {wait}s...")
                time.sleep(wait)
                continue

        except API_ERRORS as e:
            logger.error(f"Error de API no recuperable: {type(e).__name__}: {e}")
            raise

        except Exception as e:
            logger.error(f"Error inesperado: {type(e).__name__}: {e}")
            raise

    logger.error(f"Todos los reintentos fallaron. Último error: {last_error}")
    raise last_error if last_error else RuntimeError("Falló sin error registrado")

import hashlib
from PIL import Image

MAX_IMAGE_DIM = 1568
MIN_IMAGE_DIM = 100

from docx import Document  # pip install python-docx

def _process_docx(docx_bytes: bytes, name: str):
    """
    Convierte un .docx a lista de bloques: texto + tablas + imágenes.
    Las tablas se extraen por separado con python-docx para preservar estructura.
    """
    images_collected = []
    seen_hashes = set()

    def handle_image(image):
        try:
            with image.open() as img_stream:
                raw = img_stream.read()
        except Exception as e:
            logger.warning(f"No se pudo leer imagen embebida: {e}")
            return {"src": ""}

        h = hashlib.md5(raw).hexdigest()
        if h in seen_hashes:
            return {"src": ""}

        try:
            img = Image.open(io.BytesIO(raw))
            img.load()
        except Exception:
            return {"src": ""}

        if img.width < MIN_IMAGE_DIM or img.height < MIN_IMAGE_DIM:
            return {"src": ""}

        if max(img.size) > MAX_IMAGE_DIM:
            img.thumbnail((MAX_IMAGE_DIM, MAX_IMAGE_DIM), Image.LANCZOS)

        if img.mode in ("RGBA", "LA", "P"):
            img = img.convert("RGB")

        buf = io.BytesIO()
        img.save(buf, format="JPEG", quality=85, optimize=True)
        encoded = base64.standard_b64encode(buf.getvalue()).decode("utf-8")

        seen_hashes.add(h)
        images_collected.append({
            "type": "image",
            "source": {
                "type": "base64",
                "media_type": "image/jpeg",
                "data": encoded,
            }
        })
        return {"src": "", "alt": f"[IMAGEN_{len(images_collected)}]"}

    # 1) Texto con mammoth
    result = mammoth.convert_to_markdown(
        io.BytesIO(docx_bytes),
        convert_image=mammoth.images.img_element(handle_image),
    )

    import re
    text = re.sub(r'!\[(\[IMAGEN_\d+\])\]\(\)', r'\1', result.value)
    text = re.sub(r'!\[\]\(\)', '', text)

    # 2) Tablas con python-docx (preserva estructura)
    tables_md = _extract_tables_as_markdown(docx_bytes)

    # Combinar
    full_text = f"[Contenido de {name}]\n{text}"
    if tables_md:
        full_text += f"\n\n=== TABLAS DETECTADAS EN {name} ===\n{tables_md}"

    text_block = {"type": "text", "text": full_text}

    logger.info(
        f"DOCX procesado: {name} | "
        f"Texto: ~{len(text)//4:,} tokens | "
        f"Tablas: {tables_md.count('| Tabla') if tables_md else 0} | "
        f"Imágenes: {len(images_collected)}"
    )

    return [text_block] + images_collected


def _extract_tables_as_markdown(docx_bytes: bytes) -> str:
    """Extrae todas las tablas de un .docx en formato markdown."""
    try:
        doc = Document(io.BytesIO(docx_bytes))
    except Exception as e:
        logger.warning(f"No se pudo abrir docx para extraer tablas: {e}")
        return ""

    parts = []
    for i, table in enumerate(doc.tables, start=1):
        rows = []
        for row in table.rows:
            cells = [cell.text.strip().replace("\n", " ").replace("|", "/") 
                     for cell in row.cells]
            rows.append(cells)

        if not rows:
            continue

        # Filtrar filas completamente vacías
        rows = [r for r in rows if any(c for c in r)]
        if not rows:
            continue

        # Armar tabla markdown
        header = rows[0]
        body = rows[1:] if len(rows) > 1 else []

        parts.append(f"\n| Tabla {i} |")
        parts.append("| " + " | ".join(header) + " |")
        parts.append("|" + "|".join(["---"] * len(header)) + "|")
        for r in body:
            # Normalizar cantidad de columnas
            while len(r) < len(header):
                r.append("")
            parts.append("| " + " | ".join(r[:len(header)]) + " |")

    return "\n".join(parts)