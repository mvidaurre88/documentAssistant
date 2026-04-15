import base64, logging, streamlit as st
from pathlib import Path
from docx import Document
from google import genai
from google.genai import types
from openpyxl import load_workbook

logger = logging.getLogger(__name__)

# EXTENSIONES SOPORTADAS -----------------------------------------------------------------------------
SUPPORTED_EXTENSIONS = {".pdf", ".docx", ".txt", ".png", ".py", ".xlsm"}

# PROCESA LOS DIFERENTES TIPOS DE ARCHIVOS Y LOS CODIFICA PARA ENVIAR A GEMINI -----------------------
def encode_file(path: Path):

    if not path.exists():
        logger.error(f"Archivo no encontrado: {path}")
        return None

    ext = path.suffix.lower()
    if ext not in SUPPORTED_EXTENSIONS:
        logger.error(f"Tipo de archivo no soportado: '{ext}' ({path})")
        return None

    # PNG
    if ext == ".png":
        with open(path, "rb") as f:
            data = base64.standard_b64encode(f.read()).decode("utf-8")
        return types.Part.from_bytes(data=base64.b64decode(data), mime_type="image/png")

    # PDF
    if ext == ".pdf":
        with open(path, "rb") as f:
            raw = f.read()
        return types.Part.from_bytes(data=raw, mime_type="application/pdf")

    # DOCX
    if ext == ".docx":
        doc = Document(path)
        text_content = "\n".join(p.text for p in doc.paragraphs if p.text.strip())
        return types.Part.from_text(text=f"[Contenido de {path.name}]\n{text_content}")

    # TXT y PY
    if ext in (".txt", ".py"):
        text_content = path.read_text(encoding="utf-8")
        return types.Part.from_text(text=f"[Contenido de {path.name}]\n{text_content}")

    # XLSM
    if ext == ".xlsm":
        wb = load_workbook(path, read_only=True, keep_vba=True)
        lines = []
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            lines.append(f"### Hoja: {sheet_name}")
            for row in ws.iter_rows(values_only=True):
                row_text = "\t".join("" if cell is None else str(cell) for cell in row)
                if row_text.strip():
                    lines.append(row_text)
        wb.close()
        text_content = "\n".join(lines)
        return types.Part.from_text(text=f"[Contenido de {path.name}]\n{text_content}")

def send_to_gemini(promptPath: str, filePaths: list, maxTokens: int = 4096) -> str:
    
    # OBTENGO LA API KEY -----------------------------------------------------------------------------
    key = st.secrets.get("API_KEY")
    if not key:
        logger.error("API_KEY no encontrada en las variables de entorno.")
        raise EnvironmentError("API_KEY no encontrada en las variables de entorno.")

    # LEO EL PROMPT --------------------------------------------------------------------------------
    prompt_file = Path(promptPath)
    if not prompt_file.exists():
        logger.error(f"Archivo de prompt no encontrado: {promptPath}")
        raise FileNotFoundError(f"Archivo de prompt no encontrado: {promptPath}")

    prompt_text = prompt_file.read_text(encoding="utf-8").strip()
    if not prompt_text:
        logger.error(f"Archivo de prompt está vacío: {promptPath}")
        raise ValueError("El archivo de prompt está vacío.")

    # OBTENGO LOS ARCHIVOS Y LOS CODIFICO -----------------------------------------------------------
    parts = []
    for filePath in filePaths:
        file = encode_file(Path(filePath))
        if file:
            parts.append(file)

    # Prompt al final
    parts.append(types.Part.from_text(text=prompt_text))

    # Cliente y configuración
    client = genai.Client(api_key=key)

    contents = [types.Content(role="user", parts=parts)]

    generate_content_config = types.GenerateContentConfig(
        thinking_config=types.ThinkingConfig(thinking_budget=-1),
        max_output_tokens=maxTokens,
    )

    # Llamada con streaming, acumulamos la respuesta
    full_response = ""
    for chunk in client.models.generate_content_stream(
        model="gemini-2.5-flash",
        contents=contents,
        config=generate_content_config,
    ):
        if chunk.text:
            full_response += chunk.text

    return full_response