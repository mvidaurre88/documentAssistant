# IMPORTS DE PYTHON
import os
from copy import copy
from io import BytesIO
from itertools import product

# IMPORTS DE TERCEROS
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter
from PIL import ImageFont
from xlsxtpl.writerx import BookWriter

BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))

# CONSTANTES PARA CÁLCULO DE AUTOFIT
FONT_PATHS = [ (os.path.join(BASE_DIR, "fonts", "Carlito-Regular.ttf"), 1.0), ("/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf", 0.85)]
FONT_WIDTH_MULTIPLIER   = { "times new roman": 1.20, "verdana": 1.05 }
FONT_HEIGHT_MULTIPLIER  = { "times new roman": 1.15 }
DEFAULT_COLUMN_WIDTH = 8.43

def generate_TDD(context):
    
    # FORMATEO LOS DATOS OBTENIDOS
    context = expandir_tdd(context)
    
    # CARGO EL TEMPLATE
    template_path = os.path.join(BASE_DIR, "templates", "TDD.xlsx")
    writer = BookWriter(template_path)
    
    # ARMO LOS PAYLOADS PARA CADA HOJA Y LOS RENDERIZO
    payloads, cases = generate_payloads(context)
    writer.render_book(payloads=payloads)
    
    # CONFIGURACION FINAL DE HOJAS
    for ws in writer.workbook.worksheets:
        if ws.title in cases:
            distribuir_titulos(ws, cases[ws.title])
        autofit_wrapped_rows(ws)
        apply_conditional_formatting(ws)
        ws.sheet_state = "visible"
        ws.sheet_view.showGridLines = False
    writer.workbook.active = 0
    writer.custom_doc_props = ()

    # GUARDO EN BUFFER
    buffer = BytesIO()
    writer.workbook.save(buffer)
    buffer.seek(0)
    return buffer

def generate_payloads(context):
    payloads = [{
                    "tpl_idx": 0,
                    "sheet_name": "Cover",
                    "nombreBot": context.get("nombreBot", ""),
                    "codigoBot": context.get("codigoBot", ""),
                    "descripcionBot": context.get("descripcionBot", ""),
                    "version": context.get("version", ""),
                },
                {
                    "tpl_idx": 1,
                    "sheet_name": "Doc Info",
                    "cliente": context.get("cliente", ""),
                    "nombreProyecto": context.get("nombreProyecto", ""),
                    "codigoBot": context.get("codigoBot", ""),
                    "descripcionBot": context.get("descripcionBot", ""),
                    "modificaciones": context.get("modificaciones", []),
                },
                {
                    "tpl_idx": 2,
                    "sheet_name": "Descripción del proceso",
                    "inputsProceso": context.get("inputsProceso", []),
                    "pasosProceso": context.get("pasosProceso", []),
                }
            ]
    cases = {}
    for esc in context.get("escenarios", []):
        casos = esc.get("casos", [])
        is_single = len(casos) <= 1
        sheet_name = f"Escenario {esc.get('numero', '')}"

        payload = {
            "tpl_idx": 3 if is_single else 4,
            "sheet_name": sheet_name,
            "numero": esc.get("numero", ""),
            "titulo": esc.get("titulo", ""),
            "casos": casos,
            "fecha": context.get("fecha", ""),
        }

        if is_single and casos:
            payload["caso"] = casos[0]
            payload["desarrollador"] = context.get("desarrollador", "")
        elif not is_single:
            cases[sheet_name] = [c.get("titulo", "") for c in casos]

        payloads.append(payload)
    return payloads, cases

def get_column_width(ws, col_idx):
    dim = ws.column_dimensions.get(get_column_letter(col_idx))
    if dim and dim.width:
        return dim.width
    sheet_default = ws.sheet_format.defaultColWidth if ws.sheet_format else None
    return sheet_default or DEFAULT_COLUMN_WIDTH

def build_merge_map(ws):
    merge_widths = {}
    merged_skip = set()

    for merged_range in ws.merged_cells.ranges:
        min_col, min_row, max_col, max_row = merged_range.bounds
        anchor = (min_row, min_col)

        merge_widths[anchor] = sum(get_column_width(ws, col) for col in range(min_col, max_col + 1))

        celdas = set(product(range(min_row, max_row + 1),
                             range(min_col, max_col + 1)))
        merged_skip.update(celdas - {anchor})

    return merge_widths, merged_skip

def get_font_and_correction(size):
    for path, correction in FONT_PATHS:
        if os.path.exists(path):
            try:
                return ImageFont.truetype(path, int(size)), correction
            except OSError:
                continue
    return ImageFont.load_default(), 0.85

def autofit_wrapped_rows(ws, default_font_size=11, line_height_factor=1.4, min_height=18):
    PX_POR_UNIDAD = 4.5
    PADDING_CELDA = 5
    DEFAULT_FONT = "calibri"
    
    merge_widths, merged_skip = build_merge_map(ws)
    anchos_col = {
        col: dim.width or 8.43
        for col, dim in ws.column_dimensions.items()
    }
    
    def contar_lineas(texto, font, ancho_px, correction):
        total = 0
        for parrafo in texto.replace("\r\n", "\n").replace("\r", "\n").split("\n"):
            if not parrafo:
                total += 1
                continue
            
            linea_actual = ""
            lineas_parrafo = 0
            for palabra in parrafo.split(" "):
                test = f"{linea_actual} {palabra}".strip() if linea_actual else palabra
                try:
                    ancho_test = font.getlength(test) * correction
                except AttributeError:
                    ancho_test = font.getsize(test)[0] * correction
                
                if ancho_test <= ancho_px:
                    linea_actual = test
                else:
                    if linea_actual:
                        lineas_parrafo += 1
                    linea_actual = palabra
            
            if linea_actual:
                lineas_parrafo += 1
            total += max(1, lineas_parrafo)
        return total
    
    for row in ws.iter_rows():
        max_lineas = 1
        max_font = default_font_size
        max_font_name = DEFAULT_FONT
        
        for cell in row:
            if cell.value is None:
                continue
            if not (cell.alignment and cell.alignment.wrap_text):
                continue
            if (cell.row, cell.column) in merged_skip:
                continue
            
            ancho_unidades = merge_widths.get( (cell.row, cell.column), anchos_col.get(cell.column_letter, DEFAULT_COLUMN_WIDTH))
            ancho_px = max(20, ancho_unidades * PX_POR_UNIDAD - PADDING_CELDA)
            
            font_size = int(cell.font.size) if cell.font and cell.font.size else default_font_size
            font_name = cell.font.name.lower().strip() if cell.font and cell.font.name else DEFAULT_FONT
            
            font, correction = get_font_and_correction(font_size)
            effective_correction = correction * FONT_WIDTH_MULTIPLIER.get(font_name, 1.0)
            
            total_lineas = contar_lineas(str(cell.value), font, ancho_px, effective_correction)
            
            if total_lineas > max_lineas:
                max_lineas = total_lineas
                max_font = max(max_font, font_size)
                max_font_name = font_name
        
        if max_lineas > 1:
            height_mult = FONT_HEIGHT_MULTIPLIER.get(max_font_name, 1.0)
            alto = max(min_height, (max_lineas + 1) * max_font * line_height_factor * height_mult)
            ws.row_dimensions[row[0].row].height = alto

COMENTARIO_OK = "El paso se concretó como lo esperado."
COMENTARIO_FALLA = "El paso no pudo concretarse; por lo que el resultado de la prueba es correcto."
COMENTARIO_OMITIDO = "Se omitió el paso"
CAMPOS_AUXILIARES = ("excepcion", "pasoFalla", "pasosFinalesEjecutados", "datosPrueba")

def expandir_caso(caso, pasos_proceso):
    
    def _parse_paso_falla(valor):
        return int(valor) if valor not in (None, "", "null") else None
    
    def _resultado_y_comentario(n, paso_falla, finales):
        if paso_falla is None or n < paso_falla:
            return "Correcto", COMENTARIO_OK
        if n == paso_falla:
            return "Correcto", COMENTARIO_FALLA
        if n in finales:
            return "Correcto", COMENTARIO_OK
        return "N/A", COMENTARIO_OMITIDO

    paso_falla = _parse_paso_falla(caso.get("pasoFalla"))
    finales = {int(x) for x in (caso.get("pasosFinalesEjecutados") or []) if str(x).strip()}
    datos = caso.get("datosPrueba") or {}

    pasos_expandidos = []
    for paso in pasos_proceso:
        n = int(paso.get("numero"))
        resultado, comentario = _resultado_y_comentario(n, paso_falla, finales)
        # datosPrueba puede venir con clave string o int desde el JSON
        dato = datos.get(str(n)) or datos.get(n) or "N/A"

        pasos_expandidos.append({
            "numero": n,
            "descripcion": paso.get("descripcion", ""),
            "datosPrueba": dato,
            "resultado": resultado,
            "comentario": comentario,
            "captura": "N/A",
        })

    return pasos_expandidos

def expandir_tdd(context):
    pasos_proceso = context.get("pasosProceso", [])

    for escenario in context.get("escenarios", []):
        for caso in escenario.get("casos", []):
            # Si ya viene expandido (tiene "pasos"), no lo toco.
            if "pasos" in caso and caso["pasos"]:
                continue

            caso["pasos"] = expandir_caso(caso, pasos_proceso)

            # Limpio los campos auxiliares para no ensuciar el template
            for campo in CAMPOS_AUXILIARES:
                caso.pop(campo, None)

    return context

def apply_conditional_formatting(ws):
    
    # COLORES
    GREEN   = PatternFill(start_color="92D050", end_color="92D050", fill_type="solid")
    YELLOW  = PatternFill(start_color="FFFF66", end_color="FFFF66", fill_type="solid")
    RED     = PatternFill(start_color="C00000", end_color="C00000", fill_type="solid")
    
    # RANGO A FORMATEAR
    range = "D8:D200"
    
    # REGLAS DE FORMATO CONDICIONAL
    ws.conditional_formatting.add(range, CellIsRule(operator="equal", formula=['"Correcto"'], fill=GREEN))
    ws.conditional_formatting.add(range, CellIsRule(operator="equal", formula=['"Con observaciones"'], fill=YELLOW))
    ws.conditional_formatting.add(range, CellIsRule(operator="equal", formula=['"Defecto"'], fill=RED))
    
def distribuir_titulos(ws, titulos, fila_titulos=2, filas_a_replicar=(3, 4), col_inicial=2):
    
    _STYLE_ATTRS = ("font", "fill", "border", "alignment", "number_format")

    def _copiar_estilo(src, dst):
        if not src.has_style:
            return
        for attr in _STYLE_ATTRS:
            valor = getattr(src, attr)
            # number_format es un string; el resto son objetos que hay que clonar
            setattr(dst, attr, copy(valor) if attr != "number_format" else valor)
    
    if not titulos:
        return
    
    cell_titulo_base = ws.cell(row=fila_titulos, column=col_inicial)
    for i, titulo in enumerate(titulos):
        cell = ws.cell(row=fila_titulos, column=col_inicial + i)
        cell.value = titulo
        if i > 0:
            _copiar_estilo(cell_titulo_base, cell)
    
    for fila in filas_a_replicar:
        cell_base = ws.cell(row=fila, column=col_inicial)
        if cell_base.value is None:
            continue
        
        for i in range(1, len(titulos)):
            cell = ws.cell(row=fila, column=col_inicial + i)
            cell.value = cell_base.value
            _copiar_estilo(cell_base, cell)